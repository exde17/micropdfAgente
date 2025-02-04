from fastapi import FastAPI, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import io

app = FastAPI()

# Función para generar el archivo Excel
def generate_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Definir estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Determinar los datos a procesar
    if data.get("tipo") == "excel" and "proyectos" in data:
        registros = data["proyectos"]
    else:
        registros = data.get("calificaciones", [])

    if not registros:
        return b""

    # Escribir encabezados
    headers = list(registros[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Escribir datos
    for row_num, item in enumerate(registros, 2):
        for col_num, (key, value) in enumerate(item.items(), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

    # Guardar el archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()